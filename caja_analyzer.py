#!/usr/bin/env python3
"""Módulo de CAJA - Growler Garage
Responde "¿dónde está la plata?": deuda a proveedores (remitos sin pagar,
con detalle y vencidos), y movimientos de caja (retiros/deuda/inversiones)
si está cargada la planilla de Movimientos.
Genera dashboard_caja_data.json.
"""
import csv
import json
import warnings
from datetime import datetime
from pathlib import Path

warnings.simplefilter('ignore')
from openpyxl import load_workbook

PROJECT = Path('/Users/francotosto/Documents/garage_ventas')
DATA = PROJECT / 'data_gastos'
OUT = PROJECT / 'dashboard_caja_data.json'

REMITOS = {
    'GROWLER CAFE': 'MORENO - FC-Remitos - AÑO 2026 - LOCAL.csv',
    'GROWLER VIA VIEJA': 'VIA VIEJA - FC-Remitos - AÑO 2026 - LOCAL.csv',
    'COLEGIO': 'FC-Remitos - AÑO 2026 Colegio - LOCAL.csv',
    'GG Vol 2': 'GG2 - FC-Remitos - AÑO 2026  - LOCAL.csv',
    'GG Vol 4': 'GG4 - FC-Remitos - AÑO 2026.xlsx',
}
# Columnas (0-based) en la hoja LOCAL / CSV: proveedor 0, fecha 2, total 5,
# vencimiento 9, monto pago 11, pendiente 12, estado 13


def parse_num(s):
    if isinstance(s, (int, float)):
        return float(s)
    if not s or str(s).strip() in ('', '-'):
        return 0.0
    s = str(s).replace('$', '').strip()
    if ',' in s:
        s = s.replace('.', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_fecha(v):
    if isinstance(v, datetime):
        return v
    try:
        return datetime.strptime(str(v).strip(), '%d/%m/%Y')
    except (ValueError, TypeError):
        return None


def facturas_pendientes():
    """Facturas con estado pendiente/vencido/parcial por bar."""
    por_bar = {}
    detalle = []
    for local, fn in REMITOS.items():
        path = DATA / fn
        if not path.exists():
            continue
        filas = []
        if path.suffix == '.xlsx':
            wb = load_workbook(path, data_only=True)
            filas = [r for r in wb['LOCAL'].iter_rows(min_row=3, values_only=True)]
        else:
            with open(path) as f:
                r = csv.reader(f)
                next(r); next(r)
                filas = list(r)

        tot = venc_monto = 0.0
        n = n_venc = 0
        for row in filas:
            if len(row) < 14 or not row[0] or not str(row[0]).strip():
                continue
            estado = str(row[13] or '').upper().strip()
            if not ('PENDIENTE' in estado or 'VENCIDO' in estado or 'PARCIAL' in estado):
                continue
            monto = parse_num(row[12])
            if monto <= 0:
                continue
            tot += monto
            n += 1
            vto = parse_fecha(row[9])
            vencida = 'VENCIDO' in estado or (vto and vto < datetime.now())
            if vencida:
                venc_monto += monto
                n_venc += 1
            detalle.append({
                'bar': local,
                'proveedor': str(row[0]).strip()[:35],
                'fecha': parse_fecha(row[2]).strftime('%d/%m') if parse_fecha(row[2]) else '',
                'vencimiento': vto.strftime('%d/%m') if vto else '',
                'monto': round(monto),
                'vencida': bool(vencida),
            })
        if n:
            por_bar[local] = {'total': round(tot), 'facturas': n,
                              'vencido': round(venc_monto), 'vencidas': n_venc}
    detalle.sort(key=lambda x: -x['monto'])
    return por_bar, detalle[:25]


def movimientos_caja():
    """Lee la planilla de Movimientos de Caja si existe (retiros/deuda/GG3)."""
    path = DATA / 'GROWLER - Movimientos de Caja - 2026.xlsx'
    if not path.exists():
        return None
    wb = load_workbook(path, data_only=True)
    ws = wb['MOVIMIENTOS']
    movs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 4 or not row[0]:
            continue
        fecha = parse_fecha(row[0])
        monto = parse_num(row[3])
        if fecha and monto:
            movs.append({'mes': fecha.strftime('%Y-%m'), 'tipo': str(row[1] or '').strip(),
                         'quien': str(row[2] or '').strip(), 'monto': round(monto),
                         'detalle': str(row[4] or '').strip() if len(row) > 4 else ''})
    return movs


def main():
    por_bar, detalle = facturas_pendientes()
    movs = movimientos_caja()
    out = {
        'last_updated': datetime.now().isoformat(),
        'deuda_proveedores': {
            'por_bar': por_bar,
            'total': sum(b['total'] for b in por_bar.values()),
            'total_vencido': sum(b['vencido'] for b in por_bar.values()),
            'detalle_top': detalle,
        },
        'movimientos': movs,  # null = planilla no cargada todavía
    }
    with open(OUT, 'w', encoding='utf-8') as f:
        json.dump(out, f, indent=2, ensure_ascii=False)
    print(f"✓ {OUT.name} | deuda total ${out['deuda_proveedores']['total']:,.0f} "
          f"(vencido ${out['deuda_proveedores']['total_vencido']:,.0f})"
          + (" | sin planilla de movimientos" if movs is None else f" | {len(movs)} movimientos"))


if __name__ == '__main__':
    main()
