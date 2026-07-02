#!/usr/bin/env python3
"""
Sistema de Análisis de Margen - Growler Garage
Lee directamente de los archivos fuente (sin datos hardcodeados):
  - Ingresos:     sales_data.db (SQLite, alimentada por el extractor de emails)
  - Sueldos:      data_gastos/GROWLER - Liquidación Sueldos - <MES> 2026.xlsx (hoja RESUMEN LOCALES)
  - Costos Fijos: data_gastos/GROWLER - Costos Fijos - AÑO 2026.xlsx (hoja COSTOS FIJOS Y OG)
  - CMV:          data_gastos/*FC-Remitos*.csv (línea P&L = CMV)

ADM Central se distribuye: 40% CAFE, 30% VIA VIEJA, 20% COLEGIO, 5% GG2, 5% GG4.

Requiere: openpyxl (instalado en venv/). Correr con: venv/bin/python3 margin_analyzer.py
"""

import csv
import json
import sqlite3
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

PROJECT_DIR = Path(__file__).parent
DATA_GASTOS_DIR = PROJECT_DIR / 'data_gastos'
DB_PATH = PROJECT_DIR / 'sales_data.db'
OUTPUT_FILE = PROJECT_DIR / 'dashboard_margin_data.json'
CSV_LILA = PROJECT_DIR / 'margenes_analisis_lila.csv'

# Meses a reportar (agregar meses acá a medida que avanza el año)
MONTHS = ['2026-01', '2026-02', '2026-03', '2026-04', '2026-05', '2026-06']

MES_NOMBRE = {
    '2026-01': 'ENERO', '2026-02': 'FEBRERO', '2026-03': 'MARZO',
    '2026-04': 'ABRIL', '2026-05': 'MAYO', '2026-06': 'JUNIO',
    '2026-07': 'JULIO', '2026-08': 'AGOSTO', '2026-09': 'SEPTIEMBRE',
    '2026-10': 'OCTUBRE', '2026-11': 'NOVIEMBRE', '2026-12': 'DICIEMBRE',
}
NOMBRE_MES = {v: k for k, v in MES_NOMBRE.items()}
# Nombres como aparecen en la planilla de costos (capitalizados)
MES_PLANILLA = {'Enero': '2026-01', 'Febrero': '2026-02', 'Marzo': '2026-03',
                'Abril': '2026-04', 'Mayo': '2026-05', 'Junio': '2026-06',
                'Julio': '2026-07', 'Agosto': '2026-08', 'Septiembre': '2026-09',
                'Octubre': '2026-10', 'Noviembre': '2026-11', 'Diciembre': '2026-12'}

# Local en planillas -> nombre canónico (el de la BD de ventas)
LOCAL_CANONICO = {
    'MORENO': 'GROWLER CAFE',
    'VIA VIEJA': 'GROWLER VIA VIEJA',
    'COLEGIO': 'COLEGIO',
    'GG2': 'GG Vol 2',
    'GG4': 'GG Vol 4',
}
LOCALES = list(LOCAL_CANONICO.values())

# Distribución de gastos de ADM Central entre locales (definida por Franco)
ADM_DISTRIBUCION = {
    'GROWLER CAFE': 0.40,
    'GROWLER VIA VIEJA': 0.30,
    'COLEGIO': 0.20,
    'GG Vol 2': 0.05,
    'GG Vol 4': 0.05,
}

# Archivos de remitos (CMV) por local
CMV_FILES = {
    'GROWLER CAFE': 'MORENO - FC-Remitos - AÑO 2026 - LOCAL.csv',
    'GROWLER VIA VIEJA': 'VIA VIEJA - FC-Remitos - AÑO 2026 - LOCAL.csv',
    'COLEGIO': 'FC-Remitos - AÑO 2026 Colegio - LOCAL.csv',
    'GG Vol 2': 'GG2 - FC-Remitos - AÑO 2026  - LOCAL.csv',
    'GG Vol 4': 'FC-Remitos - GG4 - AÑO 2025 - GG4.csv',  # ojo: contiene 2025; 2026 pendiente
}


def parse_number(s):
    """Parsea montos en formato argentino: $ 1.234.567,89"""
    if not s or str(s).strip() in ('', '-'):
        return 0.0
    s = str(s).replace('$', '').strip()
    if ',' in s:
        s = s.replace('.', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return 0.0


def read_ingresos():
    """Ingresos por local y mes desde la BD de ventas."""
    ingresos = defaultdict(lambda: defaultdict(float))
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('SELECT date, location, SUM(total_sales) FROM sales_data GROUP BY date, location')
    for date_str, location, sales in cursor.fetchall():
        month = date_str[:7]
        if month in MONTHS:  # descarta fechas corruptas (0206-05, 2025-12, etc.)
            ingresos[location][month] += sales
    conn.close()
    return ingresos


def read_sueldos():
    """Sueldos por local y mes desde las liquidaciones Excel de Lila.

    Hoja RESUMEN LOCALES: col C = TOTAL MES $$, col D = Local, col E = Empleado.
    Se excluye personal de administración (local 'adm').
    """
    sueldos = defaultdict(lambda: defaultdict(float))
    warnings = []

    for filepath in sorted(DATA_GASTOS_DIR.glob('GROWLER - Liquidación Sueldos - * 2026.xlsx')):
        mes_nombre = filepath.stem.replace('GROWLER - Liquidación Sueldos - ', '').replace(' 2026', '').strip().upper()
        month = NOMBRE_MES.get(mes_nombre)
        if not month:
            warnings.append(f"No pude mapear mes del archivo: {filepath.name}")
            continue

        wb = load_workbook(filepath, data_only=True)
        ws = wb['RESUMEN LOCALES']
        total_mes = 0
        for row in ws.iter_rows(min_row=5, values_only=True):
            if len(row) < 5:
                continue
            monto, local_raw, empleado = row[2], row[3], row[4]
            if not (empleado and local_raw and isinstance(local_raw, str)):
                continue
            local_raw = local_raw.strip().upper()
            if local_raw not in LOCAL_CANONICO:
                continue  # excluye 'adm' y otros
            if isinstance(monto, (int, float)) and monto > 0:
                sueldos[LOCAL_CANONICO[local_raw]][month] += monto
                total_mes += monto
        if total_mes == 0:
            warnings.append(f"{mes_nombre}: liquidación sin datos (planilla vacía o sin valores calculados)")

    return sueldos, warnings


def read_costos_fijos():
    """Costos fijos por local y mes desde la planilla de Lila.

    Hoja 'COSTOS FIJOS Y OG ': col A = Mes, col D = Local, col E = Importe.
    Los gastos de ADM Central se distribuyen según ADM_DISTRIBUCION.
    """
    costos = defaultdict(lambda: defaultdict(float))
    adm_por_mes = defaultdict(float)

    filepath = DATA_GASTOS_DIR / 'GROWLER - Costos Fijos - AÑO 2026.xlsx'
    wb = load_workbook(filepath, data_only=True)
    ws = wb['COSTOS FIJOS Y OG ']

    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 5:
            continue
        mes_raw, local_raw, importe = row[0], row[3], row[4]
        month = MES_PLANILLA.get(str(mes_raw).strip()) if mes_raw else None
        if not month or month not in MONTHS:
            continue
        if not (isinstance(importe, (int, float)) and importe > 0):
            continue
        local_raw = str(local_raw).strip() if local_raw else ''
        if local_raw in LOCAL_CANONICO:
            costos[LOCAL_CANONICO[local_raw]][month] += importe
        elif local_raw == 'ADM Central':
            adm_por_mes[month] += importe

    # Distribuir ADM Central
    for month, adm_total in adm_por_mes.items():
        for local, pct in ADM_DISTRIBUCION.items():
            costos[local][month] += adm_total * pct

    return costos, dict(adm_por_mes)


def read_cmv():
    """CMV (costo de mercadería) por local y mes desde los remitos CSV."""
    cmv = defaultdict(lambda: defaultdict(float))
    warnings = []

    for local, filename in CMV_FILES.items():
        filepath = DATA_GASTOS_DIR / filename
        if not filepath.exists():
            warnings.append(f"CMV {local}: falta archivo {filename}")
            continue

        with open(filepath, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            header = next(reader, [])
            fecha_idx, total_idx, linea_idx = 2, 5, 16
            for i, col in enumerate(header):
                c = col.lower()
                if 'fecha' in c and fecha_idx == 2:
                    fecha_idx = i
                if 'total' in c and 'comprobante' in c:
                    total_idx = i
                if 'linea' in c and 'p&l' in c:
                    linea_idx = i

            count = 0
            for row in reader:
                if len(row) <= max(fecha_idx, total_idx, linea_idx):
                    continue
                try:
                    month = datetime.strptime(row[fecha_idx].strip(), '%d/%m/%Y').strftime('%Y-%m')
                except ValueError:
                    continue
                if month not in MONTHS or row[linea_idx].strip() != 'CMV':
                    continue
                amount = parse_number(row[total_idx])
                if amount > 0:
                    cmv[local][month] += amount
                    count += 1
            if count == 0:
                warnings.append(f"CMV {local}: 0 registros 2026 en {filename}")

    return cmv, warnings


def main():
    print("📊 Analizando márgenes (leyendo archivos fuente)...")

    ingresos = read_ingresos()
    sueldos, warn_sueldos = read_sueldos()
    costos, adm_por_mes = read_costos_fijos()
    cmv, warn_cmv = read_cmv()

    warnings = warn_sueldos + warn_cmv

    margin_data = {}
    for local in LOCALES:
        margin_data[local] = {}
        for month in MONTHS:
            ing = ingresos[local].get(month, 0)
            if ing == 0:
                continue
            cmv_val = cmv[local].get(month, 0)
            sueldos_val = sueldos[local].get(month, 0)
            costos_val = costos[local].get(month, 0)
            margen = ing - cmv_val - sueldos_val - costos_val
            margin_data[local][month] = {
                'ingresos': round(ing, 2),
                'cmv': round(cmv_val, 2),
                'sueldos': round(sueldos_val, 2),
                'costos_fijos': round(costos_val, 2),
                'margen': round(margen, 2),
                'margen_pct': round(margen / ing * 100, 2),
            }

    output = {
        'last_updated': datetime.now().isoformat(),
        'adm_central_por_mes': adm_por_mes,
        'adm_distribucion': ADM_DISTRIBUCION,
        'warnings': warnings,
        'margins': margin_data,
    }
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(output, f, indent=2, ensure_ascii=False)
    print(f"✓ {OUTPUT_FILE.name}")

    # CSV para Lila
    with open(CSV_LILA, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['LOCAL', 'MES', 'INGRESOS', 'CMV', 'SUELDOS', 'COSTOS_FIJOS', 'MARGEN', 'MARGEN_%'])
        for local in LOCALES:
            for month in MONTHS:
                if month not in margin_data[local]:
                    continue
                d = margin_data[local][month]
                writer.writerow([local, MES_NOMBRE[month], int(d['ingresos']), int(d['cmv']),
                                 int(d['sueldos']), int(d['costos_fijos']), int(d['margen']), d['margen_pct']])
    print(f"✓ {CSV_LILA.name}")

    print("\n" + "=" * 80)
    print("RESUMEN DE MÁRGENES")
    print("=" * 80)
    for local in LOCALES:
        print(f"\n{local}:")
        for month in MONTHS:
            if month in margin_data[local]:
                d = margin_data[local][month]
                print(f"  {month}: Margen ${d['margen']:>13,.0f} ({d['margen_pct']:>6.1f}%)")

    if warnings:
        print("\n⚠️  ADVERTENCIAS (datos incompletos):")
        for w in warnings:
            print(f"  - {w}")


if __name__ == '__main__':
    main()
