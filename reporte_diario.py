#!/usr/bin/env python3
"""Reporte diario por email - Growler Garage
Se manda todas las mañanas con: ventas de ayer vs mismo día semana pasada,
acumulado del mes, cuentas por pagar a proveedores y avisos de carga.
Destinatarios en config.secret.json -> "reporte": {"to": [...]}
"""
import json
import smtplib
import sqlite3
from datetime import date, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

PROJECT = Path('/Users/francotosto/Documents/garage_ventas')
ORDEN = ['GROWLER CAFE', 'GROWLER VIA VIEJA', 'COLEGIO', 'GG Vol 2', 'GG Vol 4']
fmt = lambda n: f"${n:,.0f}".replace(',', '.')


def datos_ventas():
    conn = sqlite3.connect(PROJECT / 'sales_data.db')
    cur = conn.cursor()
    ayer = date.today() - timedelta(days=1)
    hace8 = ayer - timedelta(days=7)
    mes_ini = ayer.replace(day=1)
    mes_ant_fin = mes_ini - timedelta(days=1)
    mes_ant_ini = mes_ant_fin.replace(day=1)
    mes_ant_corte = mes_ant_ini + (ayer - mes_ini)  # mismo día del mes anterior

    def suma(desde, hasta):
        cur.execute("SELECT location, SUM(total_sales) FROM sales_data WHERE date BETWEEN ? AND ? GROUP BY location",
                    (desde.isoformat(), hasta.isoformat()))
        return dict(cur.fetchall())

    d_ayer = suma(ayer, ayer)
    d_hace8 = suma(hace8, hace8)
    d_mes = suma(mes_ini, ayer)
    d_mes_ant = suma(mes_ant_ini, mes_ant_corte)
    conn.close()
    return ayer, d_ayer, d_hace8, d_mes, d_mes_ant


def cuentas_por_pagar():
    """Pendientes/vencidos de los remitos (snapshot en data_gastos)."""
    import warnings; warnings.simplefilter('ignore')
    from openpyxl import load_workbook
    import csv as csvmod
    pend = {}
    archivos = {
        'GROWLER CAFE': 'MORENO - FC-Remitos - AÑO 2026 - LOCAL.csv',
        'GROWLER VIA VIEJA': 'VIA VIEJA - FC-Remitos - AÑO 2026 - LOCAL.csv',
        'COLEGIO': 'FC-Remitos - AÑO 2026 Colegio - LOCAL.csv',
        'GG Vol 2': 'GG2 - FC-Remitos - AÑO 2026  - LOCAL.csv',
        'GG Vol 4': 'GG4 - FC-Remitos - AÑO 2026.xlsx',
    }
    for local, fn in archivos.items():
        path = PROJECT / 'data_gastos' / fn
        if not path.exists():
            continue
        total, vencidos = 0.0, 0
        try:
            if path.suffix == '.xlsx':
                wb = load_workbook(path, data_only=True)
                for row in wb['LOCAL'].iter_rows(min_row=3, values_only=True):
                    if len(row) > 13 and row[0]:
                        estado = str(row[13] or '').upper()
                        monto = row[12] if isinstance(row[12], (int, float)) else 0
                        if 'PENDIENTE' in estado or 'VENCIDO' in estado or 'PARCIAL' in estado:
                            total += monto
                            if 'VENCIDO' in estado:
                                vencidos += 1
            else:
                with open(path) as f:
                    r = csvmod.reader(f); next(r); next(r)
                    for row in r:
                        if len(row) > 13 and row[0].strip():
                            estado = row[13].upper()
                            monto_s = row[12].replace('$', '').replace('.', '').replace(',', '.').strip()
                            try: monto = float(monto_s)
                            except ValueError: monto = 0
                            if 'PENDIENTE' in estado or 'VENCIDO' in estado or 'PARCIAL' in estado:
                                total += monto
                                if 'VENCIDO' in estado:
                                    vencidos += 1
        except Exception:
            continue
        if total > 0:
            pend[local] = (total, vencidos)
    return pend


def avisos():
    try:
        d = json.load(open(PROJECT / 'dashboard_margin_data.json'))
        return d.get('warnings', [])
    except Exception:
        return []


def armar_html():
    ayer, d_ayer, d_hace8, d_mes, d_mes_ant = datos_ventas()
    pend = cuentas_por_pagar()
    warns = avisos()
    DIAS = ['lunes', 'martes', 'miércoles', 'jueves', 'viernes', 'sábado', 'domingo']

    tot_ayer, tot_hace8 = sum(d_ayer.values()), sum(d_hace8.values())
    tot_mes, tot_mes_ant = sum(d_mes.values()), sum(d_mes_ant.values())
    var_dia = (tot_ayer / tot_hace8 - 1) * 100 if tot_hace8 else 0
    var_mes = (tot_mes / tot_mes_ant - 1) * 100 if tot_mes_ant else 0

    filas = ''
    for loc in ORDEN:
        a, h = d_ayer.get(loc, 0), d_hace8.get(loc, 0)
        v = (a / h - 1) * 100 if h else 0
        color = '#1a7a3a' if v >= 0 else '#c02020'
        filas += (f"<tr><td style='padding:6px 10px'>{loc}</td>"
                  f"<td align='right' style='padding:6px 10px'>{fmt(a)}</td>"
                  f"<td align='right' style='padding:6px 10px;color:{color}'>{v:+.0f}%</td></tr>")

    pend_html = ''
    if pend:
        pend_html = "<h3>💳 Deuda a proveedores (remitos sin pagar)</h3><table cellspacing='0' style='border-collapse:collapse'>"
        for loc, (t, venc) in sorted(pend.items(), key=lambda x: -x[1][0]):
            extra = f" · <b style='color:#c02020'>{venc} vencidas</b>" if venc else ""
            pend_html += f"<tr><td style='padding:4px 10px'>{loc}</td><td align='right' style='padding:4px 10px'>{fmt(t)}{extra}</td></tr>"
        pend_html += f"<tr><td style='padding:4px 10px'><b>TOTAL</b></td><td align='right' style='padding:4px 10px'><b>{fmt(sum(t for t,_ in pend.values()))}</b></td></tr></table>"

    warns_html = ''
    if warns:
        warns_html = "<h3>⚠️ Avisos de carga</h3><ul>" + ''.join(f"<li>{w}</li>" for w in warns) + "</ul>"

    color_dia = '#1a7a3a' if var_dia >= 0 else '#c02020'
    color_mes = '#1a7a3a' if var_mes >= 0 else '#c02020'
    html = f"""
    <div style='font-family:-apple-system,Arial,sans-serif;max-width:560px'>
    <h2>🍺 Growler · {DIAS[ayer.weekday()]} {ayer.strftime('%d/%m')}</h2>
    <p style='font-size:1.35em;margin:4px 0'><b>{fmt(tot_ayer)}</b>
       <span style='color:{color_dia}'>({var_dia:+.0f}% vs {DIAS[ayer.weekday()]} pasado)</span></p>
    <table cellspacing='0' style='border-collapse:collapse;border-top:1px solid #ddd'>
    <tr><th align='left' style='padding:6px 10px'>Bar</th><th align='right' style='padding:6px 10px'>Ayer</th><th align='right' style='padding:6px 10px'>vs sem. pasada</th></tr>
    {filas}
    </table>
    <p><b>Mes en curso:</b> {fmt(tot_mes)} <span style='color:{color_mes}'>({var_mes:+.0f}% vs mismo período del mes pasado)</span></p>
    {pend_html}
    {warns_html}
    <p style='color:#999;font-size:0.85em'>Dashboard: colatinto.github.io/garage_ventas/dashboard_pro.html · Foto Mensual: /dashboard_margin.html</p>
    </div>"""
    asunto = f"Growler {ayer.strftime('%d/%m')}: {fmt(tot_ayer)} ({var_dia:+.0f}%)"
    return asunto, html


def enviar():
    cfg = json.load(open(PROJECT / 'config.secret.json'))
    email_cfg = cfg['email']
    destinos = cfg.get('reporte', {}).get('to', ['francototi@gmail.com'])

    asunto, html = armar_html()
    msg = MIMEMultipart('alternative')
    msg['Subject'] = asunto
    msg['From'] = f"Growler Dashboard <{email_cfg['username']}>"
    msg['To'] = ', '.join(destinos)
    msg.attach(MIMEText(html, 'html', 'utf-8'))

    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as s:
        s.login(email_cfg['username'], email_cfg['password'])
        s.sendmail(email_cfg['username'], destinos, msg.as_string())
    print(f"✓ Reporte enviado a {', '.join(destinos)}: {asunto}")


if __name__ == '__main__':
    enviar()
